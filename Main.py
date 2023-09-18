# ------------------------------------------------------
# ---------------------- main.py -----------------------
# ------------------------------------------------------

from PyQt5.QtWidgets import*
from PyQt5.uic import loadUi
from PyQt5 import QtCore, QtGui
from PyQt5.QtCore import QThread, pyqtSignal, QPropertyAnimation, QCoreApplication

#from matplotlib.backends.backend_qt5agg import (NavigationToolbar2QT as NavigationToolbar)

import time
import datetime
import numpy as np
import random
import serial
import serial.tools.list_ports; 
import openpyxl
import email, smtplib, ssl

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import ScatterChart, Reference, Series
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from page_config import FUNCTIONS


class WorkerThread(QThread):
    signal_input = pyqtSignal(float, float, bool)

    def __init__(self, parent=None):
        QThread.__init__(self)

    def run(self):
        while(1):
            self.port = [comport.device for comport in serial.tools.list_ports.comports()]
            if self.port:
                self.port_exist = True
                time.sleep(2)
                try:
                    ser= serial.Serial(str(self.port[0]), 115200, bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE)
                    print("port found", str(self.port[0]))
                    while ser.isOpen():            
                        try:
                            datastring=ser.readline().decode('ASCII')
                            values = datastring.split("&")
                            self.signal_input.emit(float(values[0]), float(values[1]), True)
                        except:
                            self.signal_input.emit(0, 0, False)
                            print("port diconected")
                            ser.close()
                        datastring = ""
                except:
                    print ("invalid port", str(self.port[0])) 
                    self.port = "" 
            else:
                self.port_exist = False
                self.signal_input.emit(0, 0, False)                

        ser.close() 


class Mainprogram(QMainWindow):
    
    def __init__(self):
        
        QMainWindow.__init__(self)

        loadUi('GUI_version1.ui',self)

        self.n_data = 60
        self.xdata = list(range(self.n_data))
        self.ydata = [0 for i in range(self.n_data)]
        self.start_count = True
        
        self.pageflag = " "
        self.data_buffer = []
        self.time_buffer = []
        self.actual_time = ""
        self.MplWidget.canvas.axes.grid()
        self.ymin = -0.5
        self.ymax = 5
        self.total_time=0
        self.auto_save_flag = False
        self.auto_save_target = 0

        self.plot_ref = None

        self.setWindowTitle("Horizon GUI version 2")

        self._translate = QtCore.QCoreApplication.translate

        self.actualPage("home")
        
        #buttons
        self.ToggleButton.clicked.connect(lambda: self.Rightmenu(200, not self.right_menu))
        self.Play_Button.clicked.connect(lambda: self.menu_buttons(not self.play))
        self.Adjust_YButton.clicked.connect(lambda: self.adjust())
        self.Home_Button.clicked.connect(lambda: self.actualPage("home"))
        self.Solar_Button.clicked.connect(lambda: self.actualPage("solar"))
        self.Wind_button.clicked.connect(lambda: self.actualPage("wind"))
        self.Hydrogen_button.clicked.connect(lambda: self.actualPage("hyd"))
        self.graph_tab.clicked.connect(lambda: self.actualPage(self.pageflag))
        self.doc_tab.clicked.connect(lambda: self.actualPage("doc"))
        self.doc_Button.clicked.connect(lambda: self.actualPage("doc"))
        self.doc_plus.clicked.connect(lambda: self.actualPage("doc"))
        self.monitor_Button.clicked.connect(lambda: self.monitor_mode(not self.monitor))
        self.Print_Button.clicked.connect(lambda: self.monitor_mode(not self.monitor, self.stackedWidget.currentIndex()))
        self.doc_plus.clicked.connect(lambda: self.actualPage("doc"))
        self.save.clicked.connect(lambda: self.save_data(0))
        self.save_as.clicked.connect(lambda: self.save_data(1))
        self.graph_tab.hide()
        self.monitor_Button.hide()
        self.doc_tab.hide()

        self.tableWidget.cellClicked.connect(self.load_data)

        #Labels   
        self.label_7.hide()

        #Frames
        self.frame_19.setMinimumSize(QtCore.QSize(0, 0))
        
        #Scrolls
        self.horizontalScrollBar.setProperty("value", 3)

        self.graph_properties('y', (20*self.horizontalScrollBar.value())+5)
        self.wt=WorkerThread()
        self.wt.start()
        self.wt.signal_input.connect(self.value)
        self.tabWidget.currentChanged.connect(self.load_combobox)
        self.auto_save_selector.currentIndexChanged.connect(self.auto_save)
        self.file_selector.currentIndexChanged.connect(self.table)
        self.horizontalScrollBar.valueChanged.connect(self.osc_range)
        self.horizontalScrollBar_2.valueChanged.connect(self.osc_volt)
        self.start_plot =time.perf_counter()
        self.play = False
        self.monitor = False
        self.right_menu = False
        self.frame_31.hide()
        self.frame_4.hide()


        self.reboot_graph()

    def graph_properties(self, linecolor, xmax):
        print(xmax)
        self.horizontalScrollBar.value()       
        self.MplWidget.canvas.axes.set_xticks(np.arange(0, xmax+round(xmax/12), round(xmax/12)))
        self.MplWidget.canvas.axes.tick_params(labelsize='small')
        self.MplWidget.canvas.axes.set_ylim(self.ymin, self.ymax)
        self.MplWidget.canvas.axes.set_xlabel('time (s)')
        self.MplWidget.canvas.axes.set_ylabel('power (w)')
        self.plot_ref.set_color(linecolor) 
        #self.MplWidget.canvas.axes.grid()
        self.MplWidget.canvas.axes.legend(('V'),loc='upper left')

    def value(self, value1, value2, port_en):
        self.v = float(value1)
        self.i = float(value2)
        self.s= self.v*self.i
        
        #print("{:.2f}".format(self.s))
        self.update_graph(port_en)
        
    def update_graph(self, port_en):

        if port_en and self.play:
            
            self.ydata = self.ydata[1:] + [self.s]
            
            self.data_buffer = self.data_buffer + [self.s]
            self.actual_time = datetime.datetime.now()                    
            self.time_buffer = self.time_buffer + [self.actual_time]

            #Voltaje y corriente actual
            self.label_4.setText(self._translate("MainWindow", "Voltaje: "+str("{:.2f}".format(self.v))))
            self.label_10.setText(self._translate("MainWindow", "Corriente: "+str("{:.2f}".format(self.i))))
            #Tiempo transcurrido
            final_time = time.perf_counter() 
            self.total_time = round(final_time-self.start_plot, 2)
            self.timeformat = time.localtime(self.total_time)
            self.label_6.setText(self._translate("MainWindow", "tiempo: "+str(self.timeformat.tm_hour-19)+":"+str(self.timeformat.tm_min)+":"+str(self.timeformat.tm_sec)))
              
            self.plot_ref.set_ydata(self.ydata)

            self.MplWidget.canvas.axes.legend(('V'),loc='upper left')
            self.MplWidget.canvas.draw()
            
            if int(self.timeformat.tm_min) >= self.auto_save_target and self.auto_save_flag:
                print("revisa el excel")
                self.save_data(0)
                self.data_buffer = []
                self.time_buffer = []
                self.menu_buttons(True)

            

        if self.eprompt_en:
            
            if self.wt.port_exist:
                
                if self.Com_Selector.count() != len(self.wt.port):

                    count_ports = len(self.wt.port)
                    for index in range(self.Com_Selector.count()):
                        self.Com_Selector.removeItem(index)

                    for index in range(count_ports):
                        self.Com_Selector.addItem("")
                        self.Com_Selector.setItemText(index, self.wt.port[index])
                    self.ePrompt(" # puertos disponibles", "green")  
                                                 
            else:
                
                for index in range(self.Com_Selector.count()):
                        self.Com_Selector.removeItem(index)
                self.menu_buttons(False)
                self.ePrompt(" # no se encontraron puertos", "red")
                             
    def reboot_graph(self):

        self.horizontalScrollBar.setProperty("value", 3)  
        self.xdata = list(range(self.n_data))
        self.ydata = [0 for i in range(self.n_data)]
        
        self.label_4.setText(self._translate("MainWindow", "Voltaje: "+str(0)))
        self.label_10.setText(self._translate("MainWindow", "Corriente: "+str(0)))
        self.label_6.setText(self._translate("MainWindow", "tiempo: "+str(0)))
        
        self.data_buffer = []
        self.time_buffer = []

        #REDRAW GRAPH
        self.MplWidget.canvas.axes.cla()
        self.MplWidget.canvas.axes.grid()
        self.MplWidget.canvas.axes.set_ylim(self.ymin, self.ymax)
        plot_refs = self.MplWidget.canvas.axes.plot(self.xdata, self.ydata)
        self.plot_ref = plot_refs[0]           
        
        self.MplWidget.canvas.draw()           

    def osc_range(self):
        
        scrollvalue = self.horizontalScrollBar.value()
        print(scrollvalue)
        if True:
            if len(self.ydata) > 20*scrollvalue:
                self.ydata = self.ydata[0:20*scrollvalue]
                
            elif len(self.ydata) < 20*scrollvalue: 
                self.ydata = self.ydata + [0 for i in range(20*scrollvalue - len(self.ydata))]
                    
            self.xdata = list(range(20*scrollvalue))
            #REDRAW GRAPH
            self.MplWidget.canvas.axes.cla()
            self.MplWidget.canvas.axes.grid()
            self.MplWidget.canvas.axes.set_ylim(self.ymin, self.ymax)            
            plot_refs = self.MplWidget.canvas.axes.plot(self.xdata, self.ydata)
            self.plot_ref = plot_refs[0]
            if self.pageflag == "solar": self.graph_properties('y', (20*self.horizontalScrollBar.value())+5)
            if self.pageflag == "wind": self.graph_properties('g', (20*self.horizontalScrollBar.value())+5)
            if self.pageflag == "hyd": self.graph_properties('b', (20*self.horizontalScrollBar.value())+5)
            self.MplWidget.canvas.draw()
            
    def osc_volt (self):

        scrollvalue = self.horizontalScrollBar_2.value()
        print (scrollvalue)
        if scrollvalue == 1: 
            self.ymin = -1 
            self.ymax = 10 
        if scrollvalue == 2: 
            self.ymin = -0.5 
            self.ymax = 5
        if scrollvalue == 3: 
            self.ymin = -0.2 
            self.ymax = 1.2

        self.MplWidget.canvas.axes.set_ylim(self.ymin, self.ymax) 
        
    def adjust(self):

        scrollvalue = self.horizontalScrollBar_2.value()
        ypos = "{:.2f}".format(self.s)           
        if scrollvalue == 1: 
            self.ymin = float(ypos)-5 
            self.ymax = float(ypos)+5
        if scrollvalue == 2: 
            self.ymin = float(ypos)-2.5 
            self.ymax = float(ypos)+2.5
        if scrollvalue == 3: 
            self.ymin = float(ypos)-0.5 
            self.ymax = float(ypos)+0.5
        self.MplWidget.canvas.axes.set_ylim(self.ymin, self.ymax)

    def Rightmenu(self, maxWidth, state):
        self.right_menu = state
        if state:
            self.frame_4.show()
            self.frame_31.hide()
            self.monitor = False
            self.frame_4.setMinimumSize(QtCore.QSize(maxWidth, 0))
        else:
            self.frame_4.hide()

    def monitor_mode(self, state, page):
        if page == 1:
            self.monitor = state
            if not state: 
                self.frame_31.hide()
            else:
                self.frame_31.show() 
                self.frame_4.hide() 
                self.right_menu = False         
                self.start_count = True
    
    def actualPage(self, page):

        #SET PAGE AND RIGHT WINDOW TEXT
              
        if page == "home":

            self.Play_Button.hide()
            self.Com_Selector.hide()
            self.label_7.hide()
            self.graph_tab.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.monitor_Button.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.doc_tab.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.stackedWidget.setCurrentWidget(self.page_home)
            self.label_solar.hide()
            self.label_wind.hide()
            self.label_hydrogen.hide()
            #self.label.setText(self._translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:600;\">  Home User Guide  </span></p></body></html>"))
            self.reboot_graph()
            self.eprompt_en = False
            self.start_count = True
            
        if page == "doc":

            self.Play_Button.hide()
            self.Com_Selector.hide()
            self.label_7.hide()
            self.graph_tab.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.monitor_Button.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.doc_tab.setStyleSheet(FUNCTIONS.toggle_page(True))
            self.stackedWidget.setCurrentWidget(self.page_doc)
            self.label.setText(self._translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:600;\">  Especifications  </span></p></body></html>"))
            self.reboot_graph()
            self.doc_tab.show()
            self.eprompt_en = False
            self.start_count = True

        if page == "solar":

            self.Play_Button.show()
            self.Com_Selector.show()
            self.graph_tab.setStyleSheet(FUNCTIONS.toggle_page(True))
            self.monitor_Button.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.doc_tab.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.label.setText(self._translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:600;\">  Solar experiments User Guide  </span></p></body></html>"))
            self.customBar('yellow')
            self.stackedWidget.setCurrentWidget(self.page_graph)
            self.pageflag = page
            self.label_solar.show()
            self.graph_tab.show()
            #Graph settings
            self.graph_properties('y', (20*self.horizontalScrollBar.value())+5)
            self.eprompt_en = True
                               
        if page == "wind":

            self.Play_Button.show()
            self.Com_Selector.show()
            self.graph_tab.setStyleSheet(FUNCTIONS.toggle_page(True))
            self.monitor_Button.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.doc_tab.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.label.setText(self._translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:600;\">  Wind experiments User Guide  </span></p></body></html>"))
            self.customBar('green')
            self.stackedWidget.setCurrentWidget(self.page_graph)
            self.pageflag = page
            self.label_wind.show()
            self.graph_tab.show()
            #Graph settings
            self.graph_properties('g', (20*self.horizontalScrollBar.value())+5)
            self.eprompt_en = True
            
        if page == "hyd":

            self.Play_Button.show()
            self.Com_Selector.show()
            self.graph_tab.setStyleSheet(FUNCTIONS.toggle_page(True))
            self.monitor_Button.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.doc_tab.setStyleSheet(FUNCTIONS.toggle_page(False))
            self.label.setText(self._translate("MainWindow", "<html><head/><body><p><span style=\" font-weight:600;\">  HYD experiments User Guide  </span></p></body></html>"))
            self.customBar('blue')
            self.stackedWidget.setCurrentWidget(self.page_graph)
            self.pageflag = page
            self.label_hydrogen.show()
            self.graph_tab.show()
            #Graph settings
            self.graph_properties('b', (20*self.horizontalScrollBar.value())+5)
            self.eprompt_en = True
                  
    def ePrompt(self, error, bcolor):

        self.label_7.setText(self._translate("MainWindow", error))
        self.frame_19.setMinimumSize(QtCore.QSize(15, 0))
        if bcolor == "red": self.frame_19.setStyleSheet("background-color: rgb(255, 82, 24);")
        if bcolor == "green": self.frame_19.setStyleSheet("background-color: rgb(140, 255, 46);")
        if bcolor == "yellow": self.frame_19.setStyleSheet("background-color: rgb(255, 243, 65);")
        self.label_7.show()
    
    def customBar(self, bcolor):

        if bcolor == "yellow": 
            self.frame_14.setStyleSheet(".QFrame{border: 1px solid yellow; border-radius: 10;}")
            self.horizontalScrollBar.setStyleSheet("")
        if bcolor == "green": self.frame_14.setStyleSheet(".QFrame{border: 1px solid green; border-radius: 10;}")
        if bcolor == "blue": self.frame_14.setStyleSheet(".QFrame{border: 1px solid blue; border-radius: 10;}")

    def menu_buttons(self, play):
        self.play = play
        if self.wt.port_exist:
            if self.play:                
                if self.start_count:
                    self.start_plot =time.perf_counter()
                    self.start_count = False
                icon = QtGui.QIcon()
                icon.addPixmap(QtGui.QPixmap("icons/pause.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
                self.Play_Button.setIcon(icon)    
                self.frame_19.setMinimumSize(QtCore.QSize(0, 0))
                self.label_7.hide()
                
            else:
                icon = QtGui.QIcon()
                icon.addPixmap(QtGui.QPixmap("icons/boton-de-play.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
                self.Play_Button.setIcon(icon)

    def auto_save(self):

        self.start_count = True
        if self.auto_save_selector.currentText() == "Desactivado":
            self.auto_save_flag = False
        elif self.auto_save_selector.currentText() == "5 minutos":
            self.auto_save_flag = True
            self.auto_save_target = 1
        elif self.auto_save_selector.currentText() == "10 minutos":
            self.auto_save_flag = True
            self.auto_save_target = 10
        elif self.auto_save_selector.currentText() == "15 minutos":
            self.auto_save_flag = True
            self.auto_save_target = 15
        elif self.auto_save_selector.currentText() == "30 minutos":
            self.auto_save_flag = True
            self.auto_save_target = 30
        elif self.auto_save_selector.currentText() == "1 hora":
            self.auto_save_flag = True
            self.auto_save_target = 59
        print(self.auto_save_target)

    def save_data(self, savemethod):
    
        
        self.start_count = True
        if self.total_time != 0:
            self.ePrompt(" # generando reporte", "yellow") 
            self.menu_buttons(False)
            if savemethod == 0:
                wb = openpyxl.load_workbook("Excel_Report.xlsx")
                s_index = len(wb.sheetnames)-2
                thin = Side(border_style="thin", color="000000")
                
                if True:
                    actual_s = wb[self.file_selector.currentText()]

                    Column = 4 
                    while(True):    
                        if actual_s.cell(4, Column).value == None:
                            break
                        else:
                            Column = Column + 2

                    actual_s.merge_cells(start_row=4, start_column=Column, end_row=4, end_column=Column+1)
                    
                    timeprint = datetime.datetime.now()
                    actual_s.cell(4, Column, "Captura:"+str(timeprint.year)+"/"+str(timeprint.month)+"/"+str(timeprint.day))

                    for i in range(4, 10, 1):

                        H = actual_s.cell(i, Column)
                        H.style = actual_s.cell(i, 2).style
                        H.value = actual_s.cell(i, 2).value
                        H.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                        S = actual_s.cell(i, Column+1)
                        S.style = actual_s.cell(i, 3).style
                        S.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    actual_s.merge_cells(start_row=5, start_column=Column, end_row=5, end_column=Column+1)
                    actual_s.merge_cells(start_row=6, start_column=Column, end_row=6, end_column=Column+1)
                    actual_s.merge_cells(start_row=7, start_column=Column, end_row=7, end_column=Column+1)
                    actual_s.merge_cells(start_row=8, start_column=Column, end_row=8, end_column=Column+1)

                    S.value = actual_s.cell(9, 3).value
                    lenght = len(self.data_buffer)
                    

                    timeprint = datetime.datetime.now()
                    actual_s.cell(4, Column, str(timeprint.year)+"/"+str(timeprint.month)+"/"+str(timeprint.day))

                    Eg = 0
                    Pp = 0
                    
                    for row in range(lenght):

                        try: 
                            b = actual_s.cell(row+10, Column, self.time_buffer[row])
                            b.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            c = actual_s.cell(row+10, Column+1, self.data_buffer[row])
                            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                            Eg = Eg + (self.data_buffer[row])*0.1
                            Pp = Pp + self.data_buffer[row]
                            
                            
                        except:
                            self.ePrompt(" # No se registran datos, capture de nuevo", "yellow")

                    self.ePrompt(" # Archivo Guardado, revisar la carpeta reporteKITHorizon", "green")
                    
                    Eg = Eg/3600
                    actual_s.cell(5, 2, round(Eg, 2)) 
                
                    actual_s.cell(6, 2, str(self.timeformat.tm_hour-19)+":"+str(self.timeformat.tm_min)+":"+str(self.timeformat.tm_sec)) 
                    
                    Pp = Pp/lenght
                    actual_s.cell(7, 2, round(Pp, 2)) 

                    actual_s.cell(8, 2, 100)
                        
                    wb.save("Excel_Report.xlsx")

                    self.sendmail()

                else:
                    self.ePrompt(" # Seleccione una hoja antes de guardar", "yellow") 

                                
            elif savemethod == 1:
                
                wb = openpyxl.load_workbook("Excel_Report.xlsx")
            
                s_index = len(wb.sheetnames)-1
                       
                formatoD = wb["FormatDaily"]
                actual_s = wb.copy_worksheet(formatoD)
                actual_s.title = self.Filename.text()   
                thin = Side(border_style="thin", color="000000")
                
                lenght = len(self.data_buffer)
                
                c = actual_s.cell(1, 1, "DATA HORIZON KIT : "+ self.Filename.text() )
                
                timeprint = datetime.datetime.now()
                actual_s.cell(4, 2, str(timeprint.year)+"/"+str(timeprint.month)+"/"+str(timeprint.day))
                
                Eg = 0
                Pp = 0
                for row in range(lenght):

                    b = actual_s.cell(row+10, 2, self.time_buffer[row])
                    b.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    c = actual_s.cell(row+10, 3, self.data_buffer[row])
                    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    Eg = Eg + (self.data_buffer[row])*0.1
                    Pp = Pp + self.data_buffer[row]
                

                Eg = Eg/3600
                actual_s.cell(5, 2, round(Eg, 2)) 

                actual_s.cell(6, 2, str(self.timeformat.tm_hour-19)+":"+str(self.timeformat.tm_min)+":"+str(self.timeformat.tm_sec)) 
                Pp = Pp/lenght
                actual_s.cell(7, 2, round(Pp, 2)) 

                actual_s.cell(8, 2, 100)
                                 
                wb.save("Excel_Report.xlsx")
                self.ePrompt(" # Archivo Guardado, revisar la carpeta reporteKITHorizon", "green")     

            
        else:
            self.ePrompt(" # Error abriendo el archivo, asegurese de no tener el archivo xlsx abierto", "red")

    def load_combobox(self):
        
        self.start_count = True
        if True:
            
            self.ePrompt(" # cargando datos", "yellow") 
            self.menu_buttons(False)
            wb = openpyxl.load_workbook("Excel_Report.xlsx")
            self.registers = wb.sheetnames
            sheets = len(self.registers)-1

            for index in range(self.file_selector.count()):
                self.file_selector.removeItem(index+1)
                
            for index in range(sheets):
                self.file_selector.addItem("")
                self.file_selector.setItemText(index+1, wb.sheetnames[index+1])
        else:
            self.ePrompt(" # No se encontro el archivo", "red") 
    
    def table(self):
        

        self.start_count = True
        if self.file_selector.currentIndex() == 0:
            self.tableWidget.clearContents()
            headers = ["Fecha(a/m/d)", "EnergiaG(w/h)", "Duracion(h:m:s)"]
            for index in range(3):

                self.tableWidget.setItem(0, index, QTableWidgetItem())
                self.tableWidget.item(0, index).setText(self._translate("MainWindow", headers[index]))

        else:
            
            self.tableWidget.clearContents()
            headers = ["Fecha(a/m/d)", "EnergiaG(w/h)", "Duracion(h:m:s)"]
            for index in range(3):
                
                self.tableWidget.setItem(0, index, QTableWidgetItem())
                self.tableWidget.item(0, index).setText(self._translate("MainWindow", headers[index]))

            wb = openpyxl.load_workbook("Excel_Report.xlsx")
            actual_s = wb[self.file_selector.currentText()] 

            Column = 2 
            while(True):    
                if actual_s.cell(4, Column).value == None:
                    datacounter = (Column-2)/2
                    break
                else:
                    Column = Column + 2

            for row in range(int(datacounter)):

                for column in range(3):
                    item = QTableWidgetItem()
                    self.tableWidget.setItem(row+1, column, item)
                    item = self.tableWidget.item(row+1, column)
                    print(actual_s.cell(column+4, 2*(row)+2).value)
                    item.setText(self._translate("MainWindow", str(actual_s.cell(column+4, 2*(row)+2).value)))
            
    def load_data (self):
        
        if self.tableWidget.currentItem()==None:
           self.ePrompt("No hay datos en esta fila", "y") 
        else:
            wb = openpyxl.load_workbook("Excel_Report.xlsx")
            actual_s = wb[self.registers[self.file_selector.currentIndex()]]
            Column = 3 + 2*(self.tableWidget.currentRow()-1) 
            self.ydata = [actual_s.cell(10, Column).value]
            row = 11
            while actual_s.cell(row, Column).value != None:
                self.ydata = self.ydata + [actual_s.cell(row, Column).value]
                row = row + 1
            self.xdata = list(range(len(self.ydata)))
            #REDRAW GRAPH
            self.MplWidget.canvas.axes.cla()
            self.MplWidget.canvas.axes.grid()
            self.MplWidget.canvas.axes.set_ylim(self.ymin, self.ymax)            
            plot_refs = self.MplWidget.canvas.axes.plot(self.xdata, self.ydata)
            self.plot_ref = plot_refs[0]
            if self.pageflag == "solar": self.graph_properties('y', len(self.ydata))
            if self.pageflag == "wind": self.graph_properties('g', len(self.ydata))
            if self.pageflag == "hyd": self.graph_properties('b', len(self.ydata))
            self.MplWidget.canvas.draw()
        
    def sendmail(self):

        subject = "Reporte automatico Monitor Energias limpias"
        sender_email = "energyprogrambot@gmail.com"
        receiver_email = "john.alexanderc131@gmail.com"
        #password = input("Type your password and press enter:")
        password = "botpw24680"
        # Create the plain-text and HTML version of your message
        text = """\
        Saludos,
        Este correo contiene el archivo adjunto con la base de datos generada de las capturas de datos de los equipos del Kit Horizon"""

        # Turn these into plain/html MIMEText objects
        part1 = MIMEText(text, "plain")

        # Create a multipart message and set headers
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = subject
        message["Bcc"] = receiver_email  # Recommended for mass emails

        # Add body to email
        message.attach(part1)

        filename = "Excel_Report.xlsx"  # In same directory as script

        # Open PDF file in binary mode
        with open(filename, "rb") as attachment:
            # Add file as application/octet-stream
            # Email client can usually download this automatically as attachment
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        # Encode file in ASCII characters to send by email    
        encoders.encode_base64(part)

        # Add header as key/value pair to attachment part
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {filename}",
        )

        # Add attachment to message and convert message to string
        message.attach(part)
        text = message.as_string()

        # Log in to server using secure context and send email
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, text)

app = QApplication([])
window = Mainprogram()
window.show()
app.exec_() 